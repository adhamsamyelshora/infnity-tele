"""
Infinity Telegram V1.0 - File & Data Managers
Handles Api_Hash.txt, Accounts.txt, Working_Accounts.txt, sessions/, tools/
"""
import os, json, csv, shutil, re
from datetime import datetime

from config import (BASE_DIR, SESSIONS_DIR, TOOLS_DIR, DATA_DIR,
                    API_HASH_FILE, ACCOUNTS_FILE, WORKING_ACCOUNTS_FILE, PROXIES_FILE)


class ApiHashManager:
    """
    Manages Api_Hash.txt file.
    Format per line: api_id,api_hash
    Example: 29907707,bc62461b45f748bc7675c81e3dc4d481
    """
    @staticmethod
    def load():
        """Load all API ID/Hash pairs from file. Returns list of (api_id, api_hash)."""
        pairs = []
        if not os.path.exists(API_HASH_FILE):
            return pairs
        with open(API_HASH_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                parts = line.split(",", 1)
                if len(parts) == 2:
                    api_id = parts[0].strip()
                    api_hash = parts[1].strip()
                    if api_id.isdigit() and len(api_hash) >= 20:
                        pairs.append((int(api_id), api_hash))
        return pairs

    @staticmethod
    def save(pairs):
        """Save API ID/Hash pairs to file."""
        with open(API_HASH_FILE, "w", encoding="utf-8") as f:
            for api_id, api_hash in pairs:
                f.write(f"{api_id},{api_hash}\n")

    @staticmethod
    def add(api_id, api_hash):
        """Add a new API pair."""
        pairs = ApiHashManager.load()
        pairs.append((int(api_id), api_hash))
        ApiHashManager.save(pairs)

    @staticmethod
    def get_for_account(index):
        """Get API pair by index (for assigning unique API per account)."""
        pairs = ApiHashManager.load()
        if not pairs:
            return None, None
        idx = index % len(pairs)
        return pairs[idx]

    @staticmethod
    def count():
        return len(ApiHashManager.load())


class AccountsManager:
    """
    Manages Accounts.txt file.
    Format: +phone,api_id@api_hash;
    Or: +phone,api_id,api_hash
    """
    @staticmethod
    def load():
        """Load accounts from file. Returns list of dicts."""
        accounts = []
        if not os.path.exists(ACCOUNTS_FILE):
            return accounts
        with open(ACCOUNTS_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip().rstrip(";")
                if not line or line.startswith("#"):
                    continue
                acc = AccountsManager._parse_line(line)
                if acc:
                    accounts.append(acc)
        return accounts

    @staticmethod
    def _parse_line(line):
        """Parse account line in multiple formats."""
        line = line.strip().rstrip(";")
        # Format: +phone,api_id@api_hash
        if "@" in line and "," in line:
            parts = line.split(",", 1)
            phone = parts[0].strip()
            rest = parts[1].strip()
            if "@" in rest:
                api_id_str, api_hash = rest.split("@", 1)
                if api_id_str.isdigit():
                    return {"phone": phone, "api_id": int(api_id_str),
                            "api_hash": api_hash.strip(), "status": "unknown"}
        # Format: +phone,api_id,api_hash
        parts = line.split(",")
        if len(parts) >= 3:
            phone = parts[0].strip()
            api_id_str = parts[1].strip()
            api_hash = parts[2].strip()
            if api_id_str.isdigit():
                return {"phone": phone, "api_id": int(api_id_str),
                        "api_hash": api_hash, "status": "unknown"}
        # Format: just phone
        if line.startswith("+") and len(line) > 5:
            return {"phone": line, "api_id": 0, "api_hash": "", "status": "no_api"}
        return None

    @staticmethod
    def save(accounts):
        """Save accounts to file."""
        with open(ACCOUNTS_FILE, "w", encoding="utf-8") as f:
            for acc in accounts:
                if acc.get("api_id") and acc.get("api_hash"):
                    f.write(f"{acc['phone']},{acc['api_id']}@{acc['api_hash']};\n")
                else:
                    f.write(f"{acc['phone']};\n")

    @staticmethod
    def add(phone, api_id=0, api_hash=""):
        """Add a new account."""
        accounts = AccountsManager.load()
        # Check duplicate
        for a in accounts:
            if a["phone"] == phone:
                a["api_id"] = api_id or a["api_id"]
                a["api_hash"] = api_hash or a["api_hash"]
                AccountsManager.save(accounts)
                return
        accounts.append({"phone": phone, "api_id": api_id, "api_hash": api_hash, "status": "unknown"})
        AccountsManager.save(accounts)

    @staticmethod
    def remove(phone):
        accounts = [a for a in AccountsManager.load() if a["phone"] != phone]
        AccountsManager.save(accounts)

    @staticmethod
    def assign_api_hashes():
        """Auto-assign API hashes from Api_Hash.txt to accounts without one."""
        accounts = AccountsManager.load()
        api_pairs = ApiHashManager.load()
        if not api_pairs:
            return 0
        changed = 0
        for i, acc in enumerate(accounts):
            if not acc.get("api_id") or not acc.get("api_hash"):
                api_id, api_hash = api_pairs[i % len(api_pairs)]
                acc["api_id"] = api_id
                acc["api_hash"] = api_hash
                changed += 1
        AccountsManager.save(accounts)
        return changed


class WorkingAccountsManager:
    """Manages Working_Accounts.txt - stores accounts confirmed as working."""

    @staticmethod
    def load():
        accounts = []
        if not os.path.exists(WORKING_ACCOUNTS_FILE):
            return accounts
        with open(WORKING_ACCOUNTS_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    accounts.append(line)
        return accounts

    @staticmethod
    def save(accounts):
        with open(WORKING_ACCOUNTS_FILE, "w", encoding="utf-8") as f:
            f.write(f"# Working Accounts - Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            for acc in accounts:
                f.write(acc + "\n")

    @staticmethod
    def add(phone):
        accounts = WorkingAccountsManager.load()
        if phone not in accounts:
            accounts.append(phone)
            WorkingAccountsManager.save(accounts)

    @staticmethod
    def clear():
        WorkingAccountsManager.save([])


class ToolsManager:
    """Manages the tools/ folder - stores scraped data."""

    @staticmethod
    def save_scraped(group_name, data_type, members):
        """Save scraped data to tools/ folder.
        data_type: 'members', 'chatters', 'admins', 'bots', 'ids'
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = re.sub(r'[^\w\-]', '_', group_name)
        filename = f"{safe_name}_{data_type}_{timestamp}.txt"
        filepath = os.path.join(TOOLS_DIR, filename)

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(f"# Scraped: {group_name} | Type: {data_type}\n")
            f.write(f"# Date: {datetime.now().isoformat()}\n")
            f.write(f"# Total: {len(members)}\n\n")
            for m in members:
                if isinstance(m, dict):
                    parts = []
                    if m.get("user_id"): parts.append(str(m["user_id"]))
                    if m.get("username"): parts.append(m["username"])
                    if m.get("name"): parts.append(m["name"])
                    if m.get("phone"): parts.append(m["phone"])
                    if m.get("access_hash"): parts.append(str(m["access_hash"]))
                    f.write(",".join(parts) + "\n")
                else:
                    f.write(str(m) + "\n")
        return filepath

    @staticmethod
    def save_ids(group_name, ids):
        """Save just IDs to tools/ folder."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = re.sub(r'[^\w\-]', '_', group_name)
        filepath = os.path.join(TOOLS_DIR, f"{safe_name}_ids_{timestamp}.txt")
        with open(filepath, "w", encoding="utf-8") as f:
            for uid in ids:
                f.write(str(uid) + "\n")
        return filepath

    @staticmethod
    def load_ids(filepath):
        """Load IDs from a tools/ file."""
        ids = []
        with open(filepath, "r") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and line.isdigit():
                    ids.append(int(line))
        return ids

    @staticmethod
    def list_files():
        """List all files in tools/ folder."""
        if not os.path.exists(TOOLS_DIR):
            return []
        return sorted(os.listdir(TOOLS_DIR), reverse=True)

    @staticmethod
    def save_excel(group_name, members):
        """Save members to Excel in tools/ folder."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = re.sub(r'[^\w\-]', '_', group_name)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            filepath = os.path.join(TOOLS_DIR, f"{safe_name}_{timestamp}.xlsx")

            wb = Workbook()
            ws = wb.active
            ws.title = "Members"

            headers = ["#", "ID", "Name", "Phone", "Username", "Status", "Access Hash", "Premium"]
            hfill = PatternFill(start_color="1a1a3e", end_color="1a1a3e", fill_type="solid")
            hfont = Font(color="FFFFFF", bold=True, size=11)

            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill, c.font, c.alignment = hfill, hfont, Alignment(horizontal="center")

            for i, m in enumerate(members, 1):
                ws.cell(row=i+1, column=1, value=i)
                ws.cell(row=i+1, column=2, value=m.get("user_id", ""))
                ws.cell(row=i+1, column=3, value=m.get("name", ""))
                ws.cell(row=i+1, column=4, value=m.get("phone", ""))
                ws.cell(row=i+1, column=5, value=m.get("username", ""))
                ws.cell(row=i+1, column=6, value=m.get("status", 0))
                ws.cell(row=i+1, column=7, value=m.get("access_hash", ""))
                ws.cell(row=i+1, column=8, value="Premium" if m.get("premium") else "Normal")

            for col in ws.columns:
                mx = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(mx + 2, 30)

            wb.save(filepath)
            return filepath
        except ImportError:
            # Fallback CSV
            filepath = os.path.join(TOOLS_DIR, f"{safe_name}_{timestamp}.csv")
            with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.DictWriter(f, fieldnames=["user_id","name","phone","username","access_hash","premium"])
                w.writeheader()
                w.writerows(members)
            return filepath


class SessionsManager:
    """Manages the sessions/ folder."""

    @staticmethod
    def list_sessions():
        return [f for f in os.listdir(SESSIONS_DIR) if f.endswith(".session")]

    @staticmethod
    def get_session_path(phone):
        clean = phone.replace("+", "").replace(" ", "")
        return os.path.join(SESSIONS_DIR, f"{clean}.session")

    @staticmethod
    def session_exists(phone):
        return os.path.exists(SessionsManager.get_session_path(phone))

    @staticmethod
    def import_sessions(source_dir):
        imported = 0
        for f in os.listdir(source_dir):
            if f.endswith(".session"):
                shutil.copy2(os.path.join(source_dir, f), os.path.join(SESSIONS_DIR, f))
                imported += 1
        return imported

    @staticmethod
    def delete_session(phone):
        path = SessionsManager.get_session_path(phone)
        if os.path.exists(path):
            os.remove(path)
            return True
        return False

    @staticmethod
    def cleanup_invalid(valid_phones):
        """Remove sessions that don't match any known account."""
        removed = 0
        valid_set = {p.replace("+","").replace(" ","") for p in valid_phones}
        for f in os.listdir(SESSIONS_DIR):
            if f.endswith(".session"):
                name = f.replace(".session", "")
                if name not in valid_set:
                    os.remove(os.path.join(SESSIONS_DIR, f))
                    removed += 1
        return removed


class ProxyManager:
    """Manages proxy list."""

    def __init__(self):
        self.proxies = self._load()
        self._idx = 0

    def _load(self):
        if os.path.exists(PROXIES_FILE):
            try:
                with open(PROXIES_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                return data if isinstance(data, list) else []
            except Exception:
                return []
        return []

    def _save(self):
        with open(PROXIES_FILE, "w") as f:
            json.dump(self.proxies, f, indent=2)

    def add(self, ptype, addr, port, user="", pwd=""):
        self.proxies.append({"type":ptype,"addr":addr,"port":int(port),
                            "user":user,"pwd":pwd,"active":True})
        self._save()

    def remove(self, idx):
        if 0 <= idx < len(self.proxies):
            self.proxies.pop(idx)
            self._save()

    def next_proxy(self):
        active = [p for p in self.proxies if p.get("active")]
        if not active: return None
        p = active[self._idx % len(active)]
        self._idx += 1
        return p

    def import_file(self, path):
        count = 0
        with open(path, "r") as f:
            for line in f:
                line = line.strip()
                if not line: continue
                if "://" in line:
                    proto, rest = line.split("://", 1)
                    u, pw, h, pt = "", "", "", "1080"
                    if "@" in rest:
                        auth, hp = rest.rsplit("@", 1)
                        if ":" in auth: u, pw = auth.split(":", 1)
                    else:
                        hp = rest
                    pp = hp.split(":")
                    h = pp[0]
                    pt = pp[1] if len(pp) > 1 else "1080"
                    self.proxies.append({"type":proto.lower(),"addr":h,"port":int(pt),
                                        "user":u,"pwd":pw,"active":True})
                    count += 1
                elif ":" in line:
                    pp = line.split(":")
                    if len(pp) >= 2:
                        self.proxies.append({"type":"socks5","addr":pp[0],"port":int(pp[1]),
                                            "user":pp[2] if len(pp)>2 else "",
                                            "pwd":pp[3] if len(pp)>3 else "","active":True})
                        count += 1
        self._save()
        return count
